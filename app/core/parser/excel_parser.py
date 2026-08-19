import logging
from pathlib import Path
from typing import Any, List, Tuple
import openpyxl

logger = logging.getLogger(__name__)


def parse_excel(file_path: Path) -> Tuple[str, List[Any], str]:
    """
    Parses an Excel spreadsheet.
    Returns: (text_summary, tables, ocr_text)
    """
    text_blocks = []
    tables = []

    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            sheet_rows = []
            MAX_SHEET_ROWS = 100
            MAX_SHEET_COLUMNS = 100
            MAX_CELL_CHARS = 2_000
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= MAX_SHEET_ROWS:
                    break
                # Filter out completely empty rows
                filtered_row = [
                    str(cell).strip()[:MAX_CELL_CHARS] if cell is not None else ""
                    for cell in row[:MAX_SHEET_COLUMNS]
                ]
                if any(filtered_row):
                    sheet_rows.append(filtered_row)

            if not sheet_rows:
                continue

            # Format as Markdown table
            headers = sheet_rows[0]
            header_line = "| " + " | ".join(headers) + " |"
            sep_line = "| " + " | ".join(["---"] * len(headers)) + " |"
            data_lines = []
            for r in sheet_rows[1:]:
                # Pad row to match header length
                padded = r + [""] * (len(headers) - len(r))
                data_lines.append("| " + " | ".join(padded[:len(headers)]) + " |")

            table_md = f"### Sheet: {sheetname}\n" + header_line + "\n" + sep_line + "\n" + "\n".join(data_lines)
            text_blocks.append(table_md)
            tables.append({
                "sheet": sheetname,
                "rows": sheet_rows[:MAX_SHEET_ROWS]
            })
        wb.close()
    except Exception as e:
        logger.error(f"Failed to parse Excel {file_path}: {e}")
        return "", [], f"Error parsing Excel: {str(e)}"

    return "\n\n".join(text_blocks), tables, ""
